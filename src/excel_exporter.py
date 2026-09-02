# src/excel_exporter.py
"""
تصدير التكاليف إلى ملف Excel
"""

import logging
from typing import Dict, Any, List
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    """مصدر التكاليف إلى Excel"""
    
    def __init__(self, cost_summary: Dict[str, Any], project_name: str = ""):
        self.cost_summary = cost_summary
        self.project_name = project_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "التكاليف"
        
        # الأنماط
        self.header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
        self.header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        self.title_font = Font(name="Arial", size=16, bold=True, color="1B4F72")
        self.subtitle_font = Font(name="Arial", size=11, color="555555")
        self.total_fill = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
        self.total_font = Font(name="Arial", size=13, bold=True, color="1B4F72")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export(self, output_path: str = None) -> str:
        if not output_path:
            reports_dir = os.path.join(os.path.dirname(__file__), '..', 'reports')
            os.makedirs(reports_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(reports_dir, f"cost_estimate_{timestamp}.xlsx")
        
        # 1. العنوان
        self.ws.merge_cells('A1:E1')
        cell = self.ws['A1']
        cell.value = "تقرير التكاليف - نظام مكافحة الحريق"
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[1].height = 35
        
        # 2. معلومات المشروع
        self.ws.merge_cells('A2:E2')
        cell = self.ws['A2']
        cell.value = f"المشروع: {self.project_name or 'غير محدد'}"
        cell.font = self.subtitle_font
        cell.alignment = Alignment(horizontal='center')
        
        self.ws.merge_cells('A3:E3')
        cell = self.ws['A3']
        cell.value = f"تاريخ الإنشاء: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        cell.font = self.subtitle_font
        cell.alignment = Alignment(horizontal='center')
        
        # 3. رأس الجدول
        row = 5
        headers = ['البند', 'الكمية', 'الوحدة', 'سعر الوحدة (ريال)', 'الإجمالي (ريال)']
        
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        self.ws.row_dimensions[row].height = 28
        
        # 4. بيانات البنود
        row += 1
        items = self.cost_summary.get('items', [])
        
        for item in items:
            self.ws.cell(row=row, column=1, value=item['item']).border = self.border
            self.ws.cell(row=row, column=2, value=item['quantity']).border = self.border
            self.ws.cell(row=row, column=3, value=item['unit']).border = self.border
            self.ws.cell(row=row, column=4, value=item['unit_price']).border = self.border
            self.ws.cell(row=row, column=5, value=item['subtotal']).border = self.border
            
            # تنسيق الأرقام
            self.ws.cell(row=row, column=4).number_format = '#,##0.00'
            self.ws.cell(row=row, column=5).number_format = '#,##0.00'
            
            # محاذاة
            self.ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            self.ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            self.ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
            
            row += 1
        
        # 5. المجاميع
        row += 1
        
        # تكلفة المواد
        self.ws.merge_cells(f'A{row}:D{row}')
        cell = self.ws.cell(row=row, column=1)
        cell.value = "📦 إجمالي تكلفة المواد"
        cell.font = self.total_font
        cell.fill = self.total_fill
        cell.border = self.border
        cell.alignment = Alignment(horizontal='right')
        
        cell = self.ws.cell(row=row, column=5)
        cell.value = self.cost_summary.get('total_material_cost', 0)
        cell.font = self.total_font
        cell.fill = self.total_fill
        cell.border = self.border
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # تكلفة التركيب
        self.ws.merge_cells(f'A{row}:D{row}')
        cell = self.ws.cell(row=row, column=1)
        cell.value = "🔧 إجمالي تكلفة التركيب"
        cell.font = self.total_font
        cell.fill = self.total_fill
        cell.border = self.border
        cell.alignment = Alignment(horizontal='right')
        
        cell = self.ws.cell(row=row, column=5)
        cell.value = self.cost_summary.get('total_labor_cost', 0)
        cell.font = self.total_font
        cell.fill = self.total_fill
        cell.border = self.border
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # الإجمالي الكلي
        self.ws.merge_cells(f'A{row}:D{row}')
        cell = self.ws.cell(row=row, column=1)
        cell.value = "💰 الإجمالي الكلي"
        cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
        cell.border = self.border
        cell.alignment = Alignment(horizontal='right')
        
        cell = self.ws.cell(row=row, column=5)
        cell.value = self.cost_summary.get('total_cost', 0)
        cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
        cell.border = self.border
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='center')
        
        # 6. ضبط عرض الأعمدة
        column_widths = [35, 15, 15, 20, 20]
        for i, width in enumerate(column_widths, 1):
            self.ws.column_dimensions[get_column_letter(i)].width = width
        
        # 7. حفظ
        Path(output_path).parent.mkdir(exist_ok=True)
        self.wb.save(output_path)
        
        logger.info(f"تم حفظ ملف Excel: {output_path}")
        return output_path